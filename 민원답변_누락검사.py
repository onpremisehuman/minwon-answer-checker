# -*- coding: utf-8 -*-
"""
국민신문고 민원답변 누락항목 검사 스크립트 v3
- 담당부서, 담당자, 연락처 누락 여부 검사
- 만족도 조사 안내 여부 검사
- 월별 통계 및 미준수 목록 시트 생성 (팀 단위 그룹화)
- 서식 정리 (외곽선, 열 너비, 색상 표시)

사용법:
1. INPUT_FILE에 검사할 엑셀 파일명 입력
2. DEPT_FILE에 부서명 목록 엑셀 파일명 입력
3. 스크립트 실행: python 민원답변_누락검사.py

출력:
- 전체목록 시트: 원본 데이터 + 누락항목/부서/담당자/연락처/만족도조사안내 열 추가
- 월별통계 시트: 답변양식 준수 현황, 만족도 조사 안내 현황, 팀별 미준수 목록
"""

import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.utils import get_column_letter
import re
import os
from datetime import datetime
from collections import defaultdict

# ============================================================
# 설정 (이 부분만 수정하면 됨)
# ============================================================
INPUT_FILE = r'국민신문고 1월 처리결과 답변.xlsx'
DEPT_FILE = r'광양시 부서명(2026.2.11.).xlsx'  # 부서명 목록 파일
VERSION = 'v3'
RESULT_COL = 6  # 처리결과 열 (F열)

# 출력 파일명 자동 생성
timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
OUTPUT_FILE = INPUT_FILE.replace('.xlsx', f'_누락검사_{timestamp}_{VERSION}.xlsx')

# ============================================================
# 스타일 정의
# ============================================================
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 1개 누락
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # 2개 누락
RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")     # 3개 누락
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================
# 부서명 목록 로드
# ============================================================
def load_dept_list():
    """부서명 엑셀에서 부서 목록 로드 (A열: 부서명)"""
    dept_set = set()
    if os.path.exists(DEPT_FILE):
        wb = openpyxl.load_workbook(DEPT_FILE)
        ws = wb.active
        for row in range(1, ws.max_row + 1):
            dept = ws.cell(row, 1).value
            if dept and isinstance(dept, str):
                dept_set.add(dept.strip())
        wb.close()
        print(f"부서명 {len(dept_set)}개 로드 완료")
    else:
        print(f"경고: 부서명 파일을 찾을 수 없습니다 - {DEPT_FILE}")
    return dept_set

DEPT_LIST = load_dept_list()

# ============================================================
# 패턴 정의
# ============================================================

# 담당자 이름 추출 패턴 (2~4글자 한글)
NAME_PATTERNS = [
    r'([가-힣]{2,4})\s*(주무관|사무관|계장|팀장|과장)',     # OOO 주무관
    r'팀\s*([가-힣]{2,4})\s*\([☎☏]?\s*[\d\s]',            # OO팀 OOO(전화번호)
    r'과\s*([가-힣]{2,4})\s*\([☎☏]?\s*[\d\s]',            # OO과 OOO(전화번호)
    r'[,、]\s*([가-힣]{2,4})\s*\)',                       # (전화번호, 이진숙)
    r'담당자\s*[:：]\s*([가-힣]{2,4})',                   # 담당자:강석호
    r'담당\s*[:：]\s*([가-힣]{2,4})',                     # 담당:강석호
    r'담당자\s+([가-힣]{2,4})\s*\(',                      # 담당자 김효천(
    r'담당\s+([가-힣]{2,4})\s*\(',                        # 담당 김효천(
    r'과\s+([가-힣]{2,4})\s*에게',                        # OO과 OOO에게
    r'팀\s+([가-힣]{2,4})\s*에게',                        # OO팀 OOO에게
]

# 이름이 아닌 단어 제외 목록
EXCLUDE_NAMES = {
    '민원', '민신문고', '민원사항', '확인', '공휴일',
}

# 전화번호 추출 패턴
TEL_PATTERNS = [
    r'[☎☏]\s*(0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4}(?:[,\s]*\d{4})*)',  # ☎061-797-2699,2868,2869
    r'(0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4})',                          # 061-797-2680
    r'\((\d{3}[-\s]?\d{4})\)',                                      # (797-2680)
    r'[☎☏]\s*(\d{3}[-\s]?\d{4})',                                  # ☎797-2680
]

# 만족도 조사 안내 패턴
SURVEY_PATTERNS = [
    '만족도조사를 실시',
    '만족도 조사를 실시',
    '만족도조사에 참여',
    '만족도 조사에 참여',
]

# ============================================================
# 추출 함수
# ============================================================
def extract_dept(content):
    """부서명 목록에서 매칭되는 부서 찾기"""
    for dept in DEPT_LIST:
        if dept in content:
            return dept
    return None

def extract_name(content):
    """담당자 이름 추출"""
    for pattern in NAME_PATTERNS:
        match = re.search(pattern, content)
        if match:
            name = match.group(1)
            if name in EXCLUDE_NAMES:
                continue
            return name
    return None

def extract_tel(content):
    """전화번호 추출"""
    for pattern in TEL_PATTERNS:
        match = re.search(pattern, content)
        if match:
            return match.group(1)
    return None

def check_survey(content):
    """만족도 조사 안내 여부 확인"""
    for pattern in SURVEY_PATTERNS:
        if pattern in content:
            return True
    return False

def apply_border_and_width(ws, max_row, max_col):
    """외곽선 적용 및 열 너비 조정"""
    column_widths = {
        'A': 20,  # 신청번호
        'B': 18,  # 신청일자
        'C': 20,  # 접수번호
        'D': 40,  # 민원제목
        'E': 15,  # 민원종류
        'F': 60,  # 처리결과
        'G': 25,  # 누락항목
        'H': 20,  # 부서
        'I': 12,  # 담당자
        'J': 25,  # 연락처
        'K': 15,  # 만족도조사안내
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).border = THIN_BORDER

# ============================================================
# 메인 처리
# ============================================================
def main():
    print("=" * 60)
    print(f"국민신문고 민원답변 누락항목 검사 {VERSION}")
    print("=" * 60)

    if not os.path.exists(INPUT_FILE):
        print(f"오류: 파일을 찾을 수 없습니다 - {INPUT_FILE}")
        return

    print(f"입력 파일: {INPUT_FILE}")
    print(f"출력 파일: {OUTPUT_FILE}")
    print()

    # 엑셀 파일 읽기
    wb = openpyxl.load_workbook(INPUT_FILE)
    ws = wb.active
    ws.title = "전체목록"

    # 기존 병합 해제
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))

    # 헤더 설정
    headers = ["신청번호", "신청일자", "접수번호", "민원제목", "민원종류",
               "처리결과", "누락항목", "부서", "담당자", "연락처", "만족도조사안내"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(1, i)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # 2행 삭제 (기존 헤더행)
    ws.delete_rows(2)

    # 카운터 초기화
    count_1 = 0  # 1개 누락
    count_2 = 0  # 2개 누락
    count_3 = 0  # 3개 누락
    total = 0
    survey_yes = 0
    survey_no = 0
    team_missing_count = defaultdict(int)

    # 데이터 처리
    for row in range(2, ws.max_row + 1):
        content = ws.cell(row, RESULT_COL).value

        if not content or str(content).strip() == "":
            continue

        content = str(content)
        total += 1

        # 만족도 조사 안내 여부 확인
        has_survey = check_survey(content)
        if has_survey:
            survey_yes += 1
        else:
            survey_no += 1

        # 정보 추출
        dept = extract_dept(content)
        name = extract_name(content)
        tel = extract_tel(content)

        # 누락 항목 확인
        missing_items = []
        if not dept:
            missing_items.append("담당부서")
        if not name:
            missing_items.append("담당자")
        if not tel:
            missing_items.append("연락처")

        # H~K열에 추출 정보 기록
        ws.cell(row, 8).value = dept if dept else ''
        ws.cell(row, 9).value = name if name else ''
        ws.cell(row, 10).value = tel if tel else ''
        ws.cell(row, 11).value = 'O' if has_survey else 'X'

        missing_count = len(missing_items)

        if missing_count > 0:
            ws.cell(row, 7).value = ", ".join(missing_items)

            if missing_count == 1:
                fill = YELLOW_FILL
                count_1 += 1
            elif missing_count == 2:
                fill = ORANGE_FILL
                count_2 += 1
            else:
                fill = RED_FILL
                count_3 += 1

            for col in range(1, 12):
                ws.cell(row, col).fill = fill

            dept_key = dept if dept else '(없음)'
            team_missing_count[dept_key] += 1
        else:
            ws.cell(row, 7).value = "없음"

    # 외곽선 및 열 너비 적용
    apply_border_and_width(ws, ws.max_row, 11)

    # 필터 설정
    ws.auto_filter.ref = f"A1:K{ws.max_row}"

    # ============================================================
    # 통계 시트 생성
    # ============================================================
    ws_stats = wb.create_sheet(title="월별통계")

    # 통계 헤더
    stats_headers = ["구분", "건수", "비율"]
    for i, header in enumerate(stats_headers, 1):
        cell = ws_stats.cell(1, i)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # 통계 계산
    total_missing = count_1 + count_2 + count_3
    compliance_rate = ((total - total_missing) / total * 100) if total > 0 else 0
    non_compliance_rate = (total_missing / total * 100) if total > 0 else 0
    survey_yes_rate = (survey_yes / total * 100) if total > 0 else 0
    survey_no_rate = (survey_no / total * 100) if total > 0 else 0

    stats_data = [
        ("전체 민원답변 수", total, "100%"),
        ("", "", ""),
        ("[ 답변양식 준수 현황 ]", "", ""),
        ("양식 준수", total - total_missing, f"{compliance_rate:.1f}%"),
        ("양식 미준수", total_missing, f"{non_compliance_rate:.1f}%"),
        ("  - 1개 누락 (노란색)", count_1, f"{(count_1/total*100) if total > 0 else 0:.1f}%"),
        ("  - 2개 누락 (주황색)", count_2, f"{(count_2/total*100) if total > 0 else 0:.1f}%"),
        ("  - 3개 누락 (빨간색)", count_3, f"{(count_3/total*100) if total > 0 else 0:.1f}%"),
        ("", "", ""),
        ("[ 만족도 조사 안내 현황 ]", "", ""),
        ("만족도 조사 안내 O", survey_yes, f"{survey_yes_rate:.1f}%"),
        ("만족도 조사 안내 X", survey_no, f"{survey_no_rate:.1f}%"),
    ]

    for row_idx, (label, count, rate) in enumerate(stats_data, 2):
        ws_stats.cell(row_idx, 1).value = label
        ws_stats.cell(row_idx, 2).value = count
        ws_stats.cell(row_idx, 3).value = rate
        for col in range(1, 4):
            ws_stats.cell(row_idx, col).border = THIN_BORDER

    # 미준수 목록 (팀 단위 그룹화)
    list_start_row = 16
    ws_stats.cell(list_start_row - 1, 1).value = "[ 답변양식 미준수 목록 (팀별 그룹화) ]"
    ws_stats.cell(list_start_row - 1, 1).font = Font(bold=True, size=12)

    list_headers = ["No", "부서(팀)", "미준수 건수"]
    for i, header in enumerate(list_headers, 1):
        cell = ws_stats.cell(list_start_row, i)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    # 미준수 목록 (건수 내림차순 정렬)
    sorted_teams = sorted(team_missing_count.items(), key=lambda x: x[1], reverse=True)
    for idx, (dept, count) in enumerate(sorted_teams, 1):
        row = list_start_row + idx
        ws_stats.cell(row, 1).value = idx
        ws_stats.cell(row, 2).value = dept
        ws_stats.cell(row, 3).value = count
        for col in range(1, 4):
            ws_stats.cell(row, col).border = THIN_BORDER

    # 열 너비 설정
    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 25
    ws_stats.column_dimensions['C'].width = 15

    # 저장
    wb.save(OUTPUT_FILE)

    # 결과 출력
    print("처리 완료!")
    print()
    print(f"전체 민원답변 수: {total}건")
    print()
    print("[ 답변양식 준수 현황 ]")
    print("-" * 40)
    print(f"양식 준수: {total - total_missing}건 ({compliance_rate:.1f}%)")
    print(f"양식 미준수: {total_missing}건 ({non_compliance_rate:.1f}%)")
    print(f"  - 1개 누락 (노란색): {count_1}건")
    print(f"  - 2개 누락 (주황색): {count_2}건")
    print(f"  - 3개 누락 (빨간색): {count_3}건")
    print()
    print("[ 만족도 조사 안내 현황 ]")
    print("-" * 40)
    print(f"만족도 조사 안내 O: {survey_yes}건 ({survey_yes_rate:.1f}%)")
    print(f"만족도 조사 안내 X: {survey_no}건 ({survey_no_rate:.1f}%)")
    print()
    print(f"저장 완료: {OUTPUT_FILE}")

if __name__ == "__main__":
    main()
