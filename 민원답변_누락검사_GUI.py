# -*- coding: utf-8 -*-
"""
국민신문고 민원답변 작성항목 점검 GUI v1.1
"""

import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import openpyxl
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
import re
import os
from datetime import datetime
from collections import defaultdict
import threading

# ============================================================
# 부서명 목록 로드
# ============================================================
def load_dept_list(dept_file):
    """A열에 부서명이 정리된 Excel 파일을 읽어 집합으로 반환한다."""
    wb = openpyxl.load_workbook(dept_file, read_only=True, data_only=True)
    try:
        ws = wb.active
        dept_list = {
            value.strip()
            for (value,) in ws.iter_rows(min_col=1, max_col=1, values_only=True)
            if (
                isinstance(value, str)
                and value.strip()
                and value.strip() not in {"부서", "부서명", "부서(팀)"}
            )
        }
    finally:
        wb.close()

    if not dept_list:
        raise ValueError("부서명 목록 파일의 A열에서 부서명을 찾지 못했습니다.")
    return dept_list

# ============================================================
# 색상 정의
# ============================================================
COLORS = {
    'header_bg': '#4ADE80',
    'header_fg': '#FFFFFF',
    'body_bg': '#F7F7F7',
    'card_bg': '#FFFFFF',
    'text_primary': '#1A1A1A',
    'text_secondary': '#6B7280',
    'accent': '#22C55E',
    'accent_hover': '#16A34A',
    'border': '#E5E7EB',
    'success': '#22C55E',
    'warning': '#F59E0B',
    'error': '#EF4444',
    'info_bg': '#FEF3C7',
    'info_border': '#F59E0B',
}

# ============================================================
# 엑셀 스타일 정의
# ============================================================
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
RED_FILL = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")

THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================
# 패턴 정의
# ============================================================
NAME_PATTERNS = [
    r'([가-힣]{2,4})\s*(주무관|사무관|계장|팀장|과장)',
    r'팀\s*([가-힣]{2,4})\s*\([☎☏]?\s*[\d\s]',
    r'과\s*([가-힣]{2,4})\s*\([☎☏]?\s*[\d\s]',
    r'[,、]\s*([가-힣]{2,4})\s*\)',
    r'담당자\s*[:：]\s*([가-힣]{2,4})',
    r'담당\s*[:：]\s*([가-힣]{2,4})',
    r'담당자\s+([가-힣]{2,4})\s*\(',
    r'담당\s+([가-힣]{2,4})\s*\(',
    r'과\s+([가-힣]{2,4})\s*에게',
    r'팀\s+([가-힣]{2,4})\s*에게',
]

EXCLUDE_NAMES = {'민원', '민신문고', '민원사항', '확인', '공휴일'}

TEL_PATTERNS = [
    r'[☎☏]\s*(0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4}(?:[,\s]*\d{4})*)',
    r'(0\d{1,2}[-\s]?\d{3,4}[-\s]?\d{4})',
    r'\((\d{3}[-\s]?\d{4})\)',
    r'[☎☏]\s*(\d{3}[-\s]?\d{4})',
]

SURVEY_PATTERNS = [
    '만족도조사를 실시',
    '만족도 조사를 실시',
    '만족도조사에 참여',
    '만족도 조사에 참여',
]

# ============================================================
# 추출 함수
# ============================================================
def extract_dept(content, dept_list):
    for dept in dept_list:
        if dept in content:
            return dept
    return None

def extract_name(content):
    for pattern in NAME_PATTERNS:
        match = re.search(pattern, content)
        if match:
            name = match.group(1)
            if name in EXCLUDE_NAMES:
                continue
            return name
    return None

def extract_tel(content):
    for pattern in TEL_PATTERNS:
        match = re.search(pattern, content)
        if match:
            return match.group(1)
    return None

def check_survey(content):
    for pattern in SURVEY_PATTERNS:
        if pattern in content:
            return True
    return False

def apply_border_and_width(ws, max_row, max_col):
    column_widths = {
        'A': 20, 'B': 18, 'C': 20, 'D': 40, 'E': 15,
        'F': 60, 'G': 25, 'H': 20, 'I': 12, 'J': 25, 'K': 15,
    }
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row, col).border = THIN_BORDER

# ============================================================
# 메인 처리 함수
# ============================================================
def process_file(input_file, output_file, dept_list, progress_callback=None, log_callback=None):
    if log_callback:
        log_callback(f"파일 로드 중: {os.path.basename(input_file)}")

    wb = openpyxl.load_workbook(input_file)
    ws = wb.active
    ws.title = "전체목록"

    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))

    headers = ["신청번호", "신청일자", "접수번호", "민원제목", "민원종류",
               "처리결과", "누락항목", "부서", "담당자", "연락처", "만족도조사안내"]
    for i, header in enumerate(headers, 1):
        cell = ws.cell(1, i)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

    ws.delete_rows(2)

    count_1 = count_2 = count_3 = total = survey_yes = survey_no = 0
    team_missing_count = defaultdict(int)
    total_rows = ws.max_row - 1

    if log_callback:
        log_callback(f"총 {total_rows}건 처리 시작...")

    for idx, row in enumerate(range(2, ws.max_row + 1), 1):
        content = ws.cell(row, 6).value

        if not content or str(content).strip() == "":
            continue

        content = str(content)
        total += 1

        has_survey = check_survey(content)
        if has_survey:
            survey_yes += 1
        else:
            survey_no += 1

        dept = extract_dept(content, dept_list)
        name = extract_name(content)
        tel = extract_tel(content)

        missing_items = []
        if not dept:
            missing_items.append("담당부서")
        if not name:
            missing_items.append("담당자")
        if not tel:
            missing_items.append("연락처")

        ws.cell(row, 8).value = dept if dept else ''
        ws.cell(row, 9).value = name if name else ''
        ws.cell(row, 10).value = tel if tel else ''
        ws.cell(row, 11).value = 'O' if has_survey else 'X'

        missing_count = len(missing_items)

        if missing_count > 0:
            ws.cell(row, 7).value = ", ".join(missing_items)
            if missing_count == 1:
                fill = YELLOW_FILL
                count_1 += 1
            elif missing_count == 2:
                fill = ORANGE_FILL
                count_2 += 1
            else:
                fill = RED_FILL
                count_3 += 1
            for col in range(1, 12):
                ws.cell(row, col).fill = fill
            dept_key = dept if dept else '(없음)'
            team_missing_count[dept_key] += 1
        else:
            ws.cell(row, 7).value = "없음"

        if progress_callback and idx % 100 == 0:
            progress_callback(int(idx / total_rows * 80))

    apply_border_and_width(ws, ws.max_row, 11)
    ws.auto_filter.ref = f"A1:K{ws.max_row}"

    if log_callback:
        log_callback("통계 시트 생성 중...")
    if progress_callback:
        progress_callback(85)

    ws_stats = wb.create_sheet(title="월별통계")

    stats_headers = ["구분", "건수", "비율"]
    for i, header in enumerate(stats_headers, 1):
        cell = ws_stats.cell(1, i)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    total_missing = count_1 + count_2 + count_3
    compliance_rate = ((total - total_missing) / total * 100) if total > 0 else 0
    non_compliance_rate = (total_missing / total * 100) if total > 0 else 0
    survey_yes_rate = (survey_yes / total * 100) if total > 0 else 0
    survey_no_rate = (survey_no / total * 100) if total > 0 else 0

    stats_data = [
        ("전체 민원답변 수", total, "100%"),
        ("", "", ""),
        ("[ 답변양식 준수 현황 ]", "", ""),
        ("양식 준수", total - total_missing, f"{compliance_rate:.1f}%"),
        ("양식 미준수", total_missing, f"{non_compliance_rate:.1f}%"),
        ("  - 1개 누락 (노란색)", count_1, f"{(count_1/total*100) if total > 0 else 0:.1f}%"),
        ("  - 2개 누락 (주황색)", count_2, f"{(count_2/total*100) if total > 0 else 0:.1f}%"),
        ("  - 3개 누락 (빨간색)", count_3, f"{(count_3/total*100) if total > 0 else 0:.1f}%"),
        ("", "", ""),
        ("[ 만족도 조사 안내 현황 ]", "", ""),
        ("만족도 조사 안내 O", survey_yes, f"{survey_yes_rate:.1f}%"),
        ("만족도 조사 안내 X", survey_no, f"{survey_no_rate:.1f}%"),
    ]

    for row_idx, (label, count, rate) in enumerate(stats_data, 2):
        ws_stats.cell(row_idx, 1).value = label
        ws_stats.cell(row_idx, 2).value = count
        ws_stats.cell(row_idx, 3).value = rate
        for col in range(1, 4):
            ws_stats.cell(row_idx, col).border = THIN_BORDER

    list_start_row = 16
    ws_stats.cell(list_start_row - 1, 1).value = "[ 답변양식 미준수 목록 (팀별 그룹화) ]"
    ws_stats.cell(list_start_row - 1, 1).font = Font(bold=True, size=12)

    list_headers = ["No", "부서(팀)", "미준수 건수"]
    for i, header in enumerate(list_headers, 1):
        cell = ws_stats.cell(list_start_row, i)
        cell.value = header
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    sorted_teams = sorted(team_missing_count.items(), key=lambda x: x[1], reverse=True)
    for idx, (dept, count) in enumerate(sorted_teams, 1):
        row = list_start_row + idx
        ws_stats.cell(row, 1).value = idx
        ws_stats.cell(row, 2).value = dept
        ws_stats.cell(row, 3).value = count
        for col in range(1, 4):
            ws_stats.cell(row, col).border = THIN_BORDER

    ws_stats.column_dimensions['A'].width = 30
    ws_stats.column_dimensions['B'].width = 25
    ws_stats.column_dimensions['C'].width = 15

    if progress_callback:
        progress_callback(95)
    if log_callback:
        log_callback("파일 저장 중...")

    wb.save(output_file)

    if progress_callback:
        progress_callback(100)

    return {
        'total': total,
        'compliance': total - total_missing,
        'compliance_rate': compliance_rate,
        'non_compliance': total_missing,
        'non_compliance_rate': non_compliance_rate,
        'count_1': count_1,
        'count_2': count_2,
        'count_3': count_3,
        'survey_yes': survey_yes,
        'survey_yes_rate': survey_yes_rate,
        'survey_no': survey_no,
        'survey_no_rate': survey_no_rate,
    }

# ============================================================
# GUI 클래스
# ============================================================
class ModernApp(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("국민신문고 민원답변 작성항목 점검")
        self.geometry("700x720")
        self.configure(bg=COLORS['body_bg'])
        self.resizable(True, True)
        self.minsize(650, 620)

        self.input_file = tk.StringVar()
        self.dept_file = tk.StringVar()

        self.create_widgets()
        self.center_window()

    def center_window(self):
        self.update_idletasks()
        x = (self.winfo_screenwidth() - 700) // 2
        y = (self.winfo_screenheight() - 720) // 2
        self.geometry(f"700x720+{x}+{y}")

    def create_widgets(self):
        # ===== 헤더 =====
        header = tk.Frame(self, bg=COLORS['header_bg'], height=50)
        header.pack(fill='x')
        header.pack_propagate(False)

        logo_frame = tk.Frame(header, bg=COLORS['header_bg'])
        logo_frame.pack(side='left', padx=15, pady=10)

        tk.Label(logo_frame, text="📋", font=('Segoe UI Emoji', 16), bg=COLORS['header_bg'], fg=COLORS['header_fg']).pack(side='left')
        tk.Label(logo_frame, text="민원답변 작성항목 점검", font=('맑은 고딕', 13, 'bold'), bg=COLORS['header_bg'], fg=COLORS['header_fg']).pack(side='left', padx=(8, 0))
        tk.Label(logo_frame, text="v1.1", font=('맑은 고딕', 8), bg='#16A34A', fg='white', padx=5, pady=1).pack(side='left', padx=(8, 0))

        # ===== 스크롤 가능한 메인 영역 =====
        container = tk.Frame(self, bg=COLORS['body_bg'])
        container.pack(fill='both', expand=True)

        canvas = tk.Canvas(container, bg=COLORS['body_bg'], highlightthickness=0)
        scrollbar = ttk.Scrollbar(container, orient='vertical', command=canvas.yview)

        self.scrollable_frame = tk.Frame(canvas, bg=COLORS['body_bg'])
        self.scrollable_frame.bind('<Configure>', lambda e: canvas.configure(scrollregion=canvas.bbox('all')))

        canvas.create_window((0, 0), window=self.scrollable_frame, anchor='nw')
        canvas.configure(yscrollcommand=scrollbar.set)

        # 마우스 휠 스크롤
        def on_mousewheel(event):
            canvas.yview_scroll(int(-1*(event.delta/120)), 'units')
        canvas.bind_all('<MouseWheel>', on_mousewheel)

        scrollbar.pack(side='right', fill='y')
        canvas.pack(side='left', fill='both', expand=True)

        # 캔버스 너비 동기화
        def configure_canvas(event):
            canvas.itemconfig(canvas.create_window((0, 0), window=self.scrollable_frame, anchor='nw'), width=event.width)
        canvas.bind('<Configure>', lambda e: canvas.itemconfig(1, width=e.width))

        main = tk.Frame(self.scrollable_frame, bg=COLORS['body_bg'])
        main.pack(fill='both', expand=True, padx=20, pady=15)

        # ===== 자료 추출방법 안내 =====
        info_card = tk.Frame(main, bg=COLORS['info_bg'], highlightbackground=COLORS['info_border'], highlightthickness=2)
        info_card.pack(fill='x', pady=(0, 10))

        tk.Label(info_card, text="⚠️ 민원답변자료 추출방법", font=('맑은 고딕', 10, 'bold'), bg=COLORS['info_bg'], fg='#92400E').pack(anchor='w', padx=12, pady=(10, 5))
        tk.Label(info_card, text="국민신문고(기관담당자) > 운영통계 > 민원자료일괄추출 > 등록\n> 신청번호, 신청일자, 접수번호, 민원제목, 민원종류, 처리결과 추출(접수일 기준)\n> 추출등록 > 신청 > 엑셀 추출 완료", font=('맑은 고딕', 9), bg=COLORS['info_bg'], fg='#78350F', justify='left').pack(anchor='w', padx=12)
        tk.Label(info_card, text="★ 추출항목이 정확히 일치해야 합니다! ★", font=('맑은 고딕', 9, 'bold'), bg=COLORS['info_bg'], fg='#DC2626').pack(anchor='w', padx=12, pady=(3, 10))

        # ===== 민원 처리결과 파일 선택 =====
        card1 = self.create_card(main)
        card1.pack(fill='x', pady=(0, 10))

        tk.Label(card1, text="1. 민원 처리결과 파일", font=('맑은 고딕', 10, 'bold'), bg=COLORS['card_bg'], fg=COLORS['text_primary']).pack(anchor='w', padx=15, pady=(12, 6))

        self.input_drop = tk.Frame(card1, bg='#F9FAFB', highlightbackground=COLORS['border'], highlightthickness=1)
        self.input_drop.pack(fill='x', padx=15, pady=(0, 12), ipady=8)

        self.input_icon = tk.Label(self.input_drop, text="📄", font=('Segoe UI Emoji', 20), bg='#F9FAFB')
        self.input_icon.pack(pady=(7, 1))
        self.input_instruction = tk.Label(self.input_drop, text="클릭하여 국민신문고 추출 파일을 선택하세요", font=('맑은 고딕', 9), bg='#F9FAFB', fg=COLORS['text_secondary'])
        self.input_instruction.pack()
        self.input_name_display = tk.Label(self.input_drop, text="", font=('맑은 고딕', 9, 'bold'), bg='#F9FAFB', fg=COLORS['accent'], wraplength=600)
        self.input_name_display.pack(pady=(2, 7))

        self.input_drop_widgets = [self.input_drop, self.input_icon, self.input_instruction, self.input_name_display]
        for w in self.input_drop_widgets:
            w.bind('<Button-1>', lambda e: self.select_input_file())
            w.bind('<Enter>', lambda e: self.set_drop_bg(self.input_drop_widgets, '#ECFDF5'))
            w.bind('<Leave>', lambda e: self.set_drop_bg(self.input_drop_widgets, '#F9FAFB'))
            w.config(cursor='hand2')

        # ===== 부서명 목록 파일 선택 =====
        dept_card = self.create_card(main)
        dept_card.pack(fill='x', pady=(0, 10))

        tk.Label(dept_card, text="2. 부서명 목록 파일", font=('맑은 고딕', 10, 'bold'), bg=COLORS['card_bg'], fg=COLORS['text_primary']).pack(anchor='w', padx=15, pady=(12, 6))

        self.dept_drop = tk.Frame(dept_card, bg='#F9FAFB', highlightbackground=COLORS['border'], highlightthickness=1)
        self.dept_drop.pack(fill='x', padx=15, pady=(0, 12), ipady=8)

        self.dept_icon = tk.Label(self.dept_drop, text="🏢", font=('Segoe UI Emoji', 20), bg='#F9FAFB')
        self.dept_icon.pack(pady=(7, 1))
        self.dept_instruction = tk.Label(self.dept_drop, text="클릭하여 A열에 부서명이 있는 Excel 파일을 선택하세요", font=('맑은 고딕', 9), bg='#F9FAFB', fg=COLORS['text_secondary'])
        self.dept_instruction.pack()
        self.dept_name_display = tk.Label(self.dept_drop, text="", font=('맑은 고딕', 9, 'bold'), bg='#F9FAFB', fg=COLORS['accent'], wraplength=600)
        self.dept_name_display.pack(pady=(2, 7))

        self.dept_drop_widgets = [self.dept_drop, self.dept_icon, self.dept_instruction, self.dept_name_display]
        for w in self.dept_drop_widgets:
            w.bind('<Button-1>', lambda e: self.select_dept_file())
            w.bind('<Enter>', lambda e: self.set_drop_bg(self.dept_drop_widgets, '#ECFDF5'))
            w.bind('<Leave>', lambda e: self.set_drop_bg(self.dept_drop_widgets, '#F9FAFB'))
            w.config(cursor='hand2')

        # ===== 검사 항목 =====
        card2 = self.create_card(main)
        card2.pack(fill='x', pady=(0, 10))

        tk.Label(card2, text="🔍 검사 항목", font=('맑은 고딕', 10, 'bold'), bg=COLORS['card_bg'], fg=COLORS['text_primary']).pack(anchor='w', padx=15, pady=(12, 6))

        items_frame = tk.Frame(card2, bg=COLORS['card_bg'])
        items_frame.pack(fill='x', padx=15, pady=(0, 12))

        for item, desc in [("담당부서", "부서명 포함"), ("담당자명", "이름 포함"), ("연락처", "전화번호 포함"), ("만족도조사", "안내 문구 포함")]:
            row = tk.Frame(items_frame, bg=COLORS['card_bg'])
            row.pack(fill='x', pady=1)
            tk.Label(row, text="✓", font=('맑은 고딕', 9), bg=COLORS['card_bg'], fg=COLORS['success']).pack(side='left')
            tk.Label(row, text=item, font=('맑은 고딕', 9, 'bold'), bg=COLORS['card_bg'], fg=COLORS['text_primary'], width=10, anchor='w').pack(side='left', padx=(5, 0))
            tk.Label(row, text=desc, font=('맑은 고딕', 9), bg=COLORS['card_bg'], fg=COLORS['text_secondary']).pack(side='left')

        # ===== 진행 상태 =====
        card3 = self.create_card(main)
        card3.pack(fill='x', pady=(0, 10))

        tk.Label(card3, text="📊 진행 상태", font=('맑은 고딕', 10, 'bold'), bg=COLORS['card_bg'], fg=COLORS['text_primary']).pack(anchor='w', padx=15, pady=(12, 6))

        self.status_label = tk.Label(card3, text="대기 중...", font=('맑은 고딕', 9), bg=COLORS['card_bg'], fg=COLORS['text_secondary'])
        self.status_label.pack(anchor='w', padx=15)

        style = ttk.Style()
        style.theme_use('clam')
        style.configure("Green.Horizontal.TProgressbar", troughcolor='#E5E7EB', background=COLORS['accent'], thickness=6)

        self.progress = ttk.Progressbar(card3, style="Green.Horizontal.TProgressbar", length=580, mode='determinate')
        self.progress.pack(padx=15, pady=(5, 3))

        self.progress_percent = tk.Label(card3, text="0%", font=('맑은 고딕', 8), bg=COLORS['card_bg'], fg=COLORS['text_secondary'])
        self.progress_percent.pack(anchor='e', padx=15, pady=(0, 10))

        # ===== 로그 =====
        log_frame = tk.Frame(main, bg='#1F2937', highlightbackground=COLORS['border'], highlightthickness=1)
        log_frame.pack(fill='x', pady=(0, 10))

        self.log_text = tk.Text(log_frame, height=4, font=('Consolas', 9), bg='#1F2937', fg='#D1D5DB', relief='flat', padx=10, pady=8, wrap='word', borderwidth=0)
        self.log_text.pack(fill='x')
        self.log_text.config(state='disabled')

        # ===== 실행 버튼 =====
        self.run_btn = tk.Button(main, text="🚀  검사 시작", font=('맑은 고딕', 11, 'bold'), bg=COLORS['accent'], fg='white', relief='flat', cursor='hand2', activebackground=COLORS['accent_hover'], activeforeground='white', command=self.run_analysis, state='disabled', borderwidth=0)
        self.run_btn.pack(fill='x', ipady=12)

        self.run_btn.bind('<Enter>', lambda e: self.run_btn.config(bg=COLORS['accent_hover']) if self.run_btn['state'] == 'normal' else None)
        self.run_btn.bind('<Leave>', lambda e: self.run_btn.config(bg=COLORS['accent']) if self.run_btn['state'] == 'normal' else None)

    def create_card(self, parent):
        return tk.Frame(parent, bg=COLORS['card_bg'], highlightbackground=COLORS['border'], highlightthickness=1)

    def set_drop_bg(self, widgets, color):
        for widget in widgets:
            widget.config(bg=color)

    def update_run_button_state(self):
        state = 'normal' if self.input_file.get() and self.dept_file.get() else 'disabled'
        self.run_btn.config(state=state)

    def select_input_file(self):
        file_path = filedialog.askopenfilename(title="민원 처리결과 엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path:
            self.input_file.set(file_path)
            self.input_instruction.config(text="선택된 파일:")
            self.input_name_display.config(text=os.path.basename(file_path))
            self.input_icon.config(text="📊")
            self.update_run_button_state()
            self.log(f"민원 파일 선택: {os.path.basename(file_path)}")

    def select_dept_file(self):
        file_path = filedialog.askopenfilename(title="부서명 목록 엑셀 파일 선택", filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")])
        if file_path:
            self.dept_file.set(file_path)
            self.dept_instruction.config(text="선택된 파일:")
            self.dept_name_display.config(text=os.path.basename(file_path))
            self.dept_icon.config(text="✅")
            self.update_run_button_state()
            self.log(f"부서명 파일 선택: {os.path.basename(file_path)}")

    def log(self, message):
        self.log_text.config(state='normal')
        self.log_text.insert('end', f"[{datetime.now().strftime('%H:%M:%S')}] {message}\n")
        self.log_text.see('end')
        self.log_text.config(state='disabled')

    def update_progress(self, value):
        self.progress['value'] = value
        self.progress_percent.config(text=f"{value}%")
        self.update_idletasks()

    def run_analysis(self):
        if not self.input_file.get() or not self.dept_file.get():
            messagebox.showwarning("알림", "민원 처리결과 파일과 부서명 목록 파일을 모두 선택해주세요.")
            return

        input_path = self.input_file.get()
        dept_path = self.dept_file.get()

        # outputs 폴더 생성
        input_dir = os.path.dirname(input_path)
        outputs_dir = os.path.join(input_dir, 'outputs')
        if not os.path.exists(outputs_dir):
            os.makedirs(outputs_dir)

        # 출력파일 경로 (outputs 폴더 내)
        input_name = os.path.basename(input_path)
        output_name = input_name.replace('.xlsx', f'_누락검사_{datetime.now().strftime("%Y%m%d_%H%M%S")}_v1.1.xlsx')
        output_path = os.path.join(outputs_dir, output_name)

        self.run_btn.config(state='disabled', text='⏳ 검사 중...', bg='#9CA3AF')
        self.status_label.config(text="분석 진행 중...", fg=COLORS['text_secondary'])

        def process():
            try:
                dept_list = load_dept_list(dept_path)
                self.after(0, lambda: self.log(f"부서명 {len(dept_list)}개 로드 완료"))
                result = process_file(input_path, output_path, dept_list,
                    progress_callback=lambda v: self.after(0, lambda: self.update_progress(v)),
                    log_callback=lambda m: self.after(0, lambda: self.log(m)))
                self.after(0, lambda: self.on_complete(result, output_path))
            except Exception as e:
                self.after(0, lambda: self.on_error(str(e)))

        threading.Thread(target=process, daemon=True).start()

    def on_complete(self, result, output_path):
        self.status_label.config(text="✅ 완료!", fg=COLORS['success'])
        self.run_btn.config(state='normal', text='🚀  검사 시작', bg=COLORS['accent'])

        self.log("=" * 40)
        self.log(f"전체: {result['total']}건 | 준수: {result['compliance']}건({result['compliance_rate']:.1f}%) | 미준수: {result['non_compliance']}건")
        self.log(f"만족도조사 O: {result['survey_yes']}건 | X: {result['survey_no']}건")
        self.log(f"저장: {os.path.basename(output_path)}")

        msg = f"""검사 완료!

📊 전체: {result['total']}건
✅ 양식 준수: {result['compliance']}건 ({result['compliance_rate']:.1f}%)
❌ 양식 미준수: {result['non_compliance']}건 ({result['non_compliance_rate']:.1f}%)
⭕ 만족도조사 안내 O: {result['survey_yes']}건 ({result['survey_yes_rate']:.1f}%)

결과 파일을 열어볼까요?"""

        if messagebox.askyesno("완료", msg):
            os.startfile(output_path)

    def on_error(self, error_msg):
        self.status_label.config(text="❌ 오류", fg=COLORS['error'])
        self.run_btn.config(state='normal', text='🚀  검사 시작', bg=COLORS['accent'])
        self.log(f"오류: {error_msg}")
        messagebox.showerror("오류", f"처리 중 오류:\n{error_msg}")

if __name__ == "__main__":
    app = ModernApp()
    app.mainloop()
